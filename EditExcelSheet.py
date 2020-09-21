from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd 
import os


def EditExceelSheet():
    #
    #Start by opening the spreadsheet and selecting the main sheet
    workbook = load_workbook(filename=input("Enter sheet Name : "))
    sheet = workbook.active
    edit=input("Enter Data For cell : ")
    
    sheet.cell(row=int(input("Enter Row No : ")), column=int(input("Enter column No  :  "))).value =edit
    
    # Save the spreadsheet
    workbook.save(filename=input("Enter sheet Name to save : "))

EditExceelSheet()